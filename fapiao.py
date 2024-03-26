#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author: qi xy 
# date: 2024/03/22    time: 17:15 
# filename：fapiao.py

import os
import zipfile
import shutil
import pdfplumber
import re
from openpyxl import Workbook, load_workbook


def extract_zip_file(file_path, output_directory):
    with zipfile.ZipFile(file_path, 'r') as zf:
        zf.extractall(output_directory)


def extract_table_from_pdf(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        tables = first_page.extract_tables()
    return tables


def extract_text_from_pdf(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()
    return text


def filter_list_with_prefix(lst,prefix):
    filtered_list = [idx for idx in lst if idx.startwith(prefix)]
    return filtered_list


# 获得二维列表某个值的一维索引值的另一种方法
def getTwoDimensionListIndex(*args):
    """获得二维列表某个值的一维索引值的另一种方法"""
    idx1 =0
    idx2 =0
    idx3 =0
    for i in range(len(args[0])):
        for j in range(len(args[0][i])):
            for k in range(len(args[0][i][j])):
                if args[0][i][j][k] is None:
                    continue
                elif len(args) == 3 and args[1] in args[0][i][j][k]:
                    idx1,idx2 = i,j

                elif len(args) == 4 and (args[1] in args[0][i][j][k] or args[2] in args[0][i][j][k]):
                    idx1, idx2= i, j

    for kk in range(len(args[0][idx1][idx2])):
        if args[0][idx1][idx2][kk] is None:
            continue
        elif len(args) == 3 and args[2] in args[0][idx1][idx2][kk]:
            idx3 = kk
        elif len(args) == 4 and args[3] in args[0][idx1][idx2][kk]:
            idx3 = kk

    return idx1, idx2, idx3


# 输入目录和文件名
in_directory = 'E:\\SIAT\\baoxiao2\\fapiao'
in_name = '滴滴出行行程报销单.pdf'

# 输出目录和文件名
out_directory = 'E:\\SIAT\\baoxiao2\\output_fapiao'
output_excelName = 'new_pdf-wjh.xlsx'
file_name_out = os.path.join(out_directory, output_excelName)

# 创建输出目录
os.makedirs(out_directory, exist_ok=True)

# 加载现有的Excel文件或创建新的文件
if not os.path.exists(file_name_out):
    wb = Workbook()
    wb.save(file_name_out)
workbook = load_workbook(filename=file_name_out)
sheet = workbook.active

row_index = 1

# 遍历文件
for file_name in os.listdir(in_directory):
    file_path = os.path.join(in_directory, file_name)

    if os.path.isfile(file_path):
        print(file_name)

    if file_name.endswith('.zip'):
        # 解压缩ZIP文件
        extract_directory = os.path.join(out_directory, 'temp')
        os.makedirs(extract_directory, exist_ok=True)
        extract_zip_file(file_path, extract_directory)

        # 处理解压后的文件
        extracted_file = os.path.join(extract_directory, in_name)
        tables = extract_table_from_pdf(extracted_file)

        for table in tables:
            for row in table:
                for col_index, cell in enumerate(row):
                    sheet.cell(row=row_index, column=col_index + 1, value=cell)
                row_index += 1

        # 清理临时文件夹
        shutil.rmtree(extract_directory)

    if file_name.endswith('.pdf'):
        # 处理PDF文件
        pdf_file = file_path
        with pdfplumber.open(pdf_file) as p1:
            for page in range(len(p1.pages)):
                #page = 29
                page1 = p1.pages[page]
                print('page'+str(page+1))
                #tables = extract_table_from_pdf(page1)
                tables = page1.extract_tables()
                texts = page1.extract_text()
                text1 = texts.splitlines()

                string_mat = ['合^\s+\s$计']
                text_price = []
                #text = extract_text_from_pdf(page1)
                #a = tables[0][3][1]
                if len(tables)>0:
                    idx11,idx22,idx33 = getTwoDimensionListIndex(tables, '销','名')
                    supply_name = tables[idx11][idx22][idx33].splitlines()
                    supply_name1 = supply_name[0]
                else:
                    #for i in range(len(text1)):
                        #text_price.append(re.findall(r'[合^\s+\s$计]', text1[i]))
                        #text_price.append(re.findall(r'[合^\s+\s$计]', text1[i]))
                    supply_name = [idx for idx in text1 if '名' in idx]
                    supply_name1 = supply_name[2]
                '''
                filter_money = [idx for idx in text1 if '价税合计' in idx]
                # filter_data = filter_list_with_prefix(text1,'开票日期')
                money = re.findall(r"\d+\.?\d*", filter_money[0])
                '''
                #if len(money)==0:
                '''
                idx111, idx222, idx333 = getTwoDimensionListIndex(tables, '货物' or '项目', '价税合计', '小写')
                total_price = re.findall(r"\d+\.?\d*", tables[idx111][idx222][idx333])
                total_price1 = total_price[0]
                # tax = tables[0][1][10].splitlines()
                # tax1 = tax[1]
                    '''
                #idx1111, idx2222, idx3333 = getTwoDimensionListIndex(tables, '货物' or '项目', '规格型号', '税率')
                #table1 = tables[idx1111][idx2222][:]
                length_tables = len(tables)
                table1 = []
                price_tax_list=[]
                if length_tables>0:
                    for i in range(length_tables):
                        if len(tables[i])>0:
                            for j in range(len(tables[i])):
                                if len(tables[i][j])>0:
                                   table1.append(list(filter(None, tables[i][j])))

                    for i0 in range(len(table1)):
                        if len(table1[i])>0:
                            for j0 in range(len(table1[i0])):
                                if len(table1[i0][j0])>0:
                                    if '¥' in table1[i0][j0]:
                                        price_tax_list.append(table1[i0][j0])
                    if len(price_tax_list)>2:
                        price_index = price_tax_list[0].find("¥")
                        price = price_tax_list[0][price_index + 1:len(price_tax_list[0])]
                        tax_index = price_tax_list[1].find("¥")
                        tax1 = price_tax_list[1][tax_index + 1:len(price_tax_list[1])]
                        total_price_index = price_tax_list[2].find("¥")
                        total_price1 = price_tax_list[2][total_price_index + 1:len(price_tax_list[2])]
                    else:
                        tax_string_split = price_tax_list[0].splitlines()
                        price_tax_list1 = [idx for idx in tax_string_split if '¥' in idx]
                        price_tax = re.findall(r"\d+\.?\d*", price_tax_list1[0])
                        price = price_tax[0]
                        tax1 = price_tax[1]
                        total_price_index = price_tax_list[1].find("¥")
                        total_price1 = price_tax_list[1][total_price_index + 1:len(price_tax_list[1])]
                else:
                    text_price = [idx for idx in text1 if '合' in idx]
                    price_tax = re.findall(r"\d+\.?\d*", text_price[0])
                    price = price_tax[0]
                    tax1 = price_tax[1]
                    total_price_index = text_price[1].find("¥")
                    total_price1 = text_price[1][total_price_index + 1:len(text_price[1])]
                '''
                price_tax_list = [s for s in table1 if '¥' in s]
                if idx3333 > 0:
                    #price_tax_list = [s for s in table1 if '¥' in s]
                    price_index = price_tax_list[0].find("¥")
                    price = price_tax_list[0][price_index+1:len(price_tax_list[0])]
                    tax_index = price_tax_list[1].find("¥")
                    tax1 = price_tax_list[1][tax_index+1:len(price_tax_list[1])]
                else:
                    tax_string = tables[idx1111][idx2222][idx3333]
                    tax_string_split = tax_string.splitlines()
                    price_tax_list = [idx for idx in tax_string_split if '¥' in idx]
                    price_tax = re.findall(r"\d+\.?\d*", price_tax_list[0])
                    price = price_tax[0]
                    tax1 = price_tax[1]
                    '''
                '''
                tax_index = tax_string.find("%")
                tax = tax_string[tax_index-2:tax_index]
                if '*' in tax:
                    tax1 = '***'
                else:
                    tax10 = re.findall(r"\d+\.?\d*", tax)
                    a1 = float(tax10[0])/100
                    a2 = float(total_price1)/(1+a1)
                    tax1 = round(a2*a1,2)
                    tax1 = str(tax1)
                    '''
                '''
                if len(tax) == 0:
                    tax1 = '***'
                else:
                    tax1 = tax[0]
                    
                
                elif len(money)== 1:
                    tax1 = '***'
                    total_price1 = money[0]
                elif len(money)==2:
                    tax1 = money[1]
                    total_price1 = money[0]
                    '''
                #filter_data=[idx for idx in text1 if idx.startswith('开票日期')]
                filter_data = [idx for idx in text1 if '开票' in idx]
                # date_index = text1.index('开票日期')
                invoice_date = re.findall(r"\d+\.?\d*",filter_data[0])
                invoice_date1 = invoice_date[0] + '-' + invoice_date[1] + '-' + invoice_date[2]
                filter_invoice_num = [idx for idx in text1 if '发票号码'in idx]
                # filter_invoice_num = filter_list_with_prefix(text1, '发票号码')
                if len(filter_invoice_num)>0:
                    invoice_num = re.findall(r"\d+\.?\d*",filter_invoice_num[0])
                    invoice_num1=invoice_num[0]
                else:
                    invoice_num1 = '*****'
                filter_invoice_type = [idx for idx in text1 if '增值税' in idx or '普' in idx]
                if len(filter_invoice_type)>0:
                    if '增值税' in filter_invoice_type[0]:
                        invoice_type = '增值税发票'
                    elif '普' in filter_invoice_type[0]:
                        invoice_type = '普通发票'
                else:
                    invoice_type = '****'
                title_row1 = ['供应商','发票号码','发票时间','发票金额','发票税额','发票类型']
                data_row2 = [supply_name1,invoice_num1,invoice_date1,total_price1,tax1,invoice_type]
                print(type(invoice_num1))

                #sheet.cell(row=1, column=col_index + 1, value=cell)

                # 处理提取的表格数据
                if row_index == 1:
                    for column_1 in range(len(title_row1)):
                        sheet.cell(row_index, column_1 + 1, title_row1[column_1])
                        # sheet.cell(row=i, column=j).value = '要修改成的数值'
                        #sheet.cell(row_index+1, column_1 + 1, data_row2[column_1])
                row_index = row_index + 1
                for column_2 in range(len(data_row2)):
                    sheet.cell(row_index, column_2 + 1, data_row2[column_2])
                # 处理提取的文本数据
                #text_lines = text.split('\n')
                # 处理text_lines中的数据并写入Excel

# 保存Excel文件
workbook.save(filename=file_name_out)